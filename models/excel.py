__author__ = 'Robert W. Curtiss'
__project__ = 'Breathing Air Solutions'

"""
====================================================
Author: Robert W. Curtiss
    Project: Breathing Air Solutions
    File: excel
    Created: Apr, 15, 2020
    
    Description:
    
===================================================
"""
from openpyxl import load_workbook, Workbook


class MyExcel(object):
    def __init__(self):
        self.columns = []
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.current_row_number = 1

    def init(self, title, header, rows):
        #workbook = load_workbook(file_name + r'.xlsx')
        self.sheet.title = title
        i = 1
        x = self.sheet.active_cell
        y = self.sheet.selected_cell
        z = self.sheet.columns
        self.sheet.insert_cols(1,len(header)-1)
        self.sheet.insert_rows(idx=1, amount=rows)
        self.current_row_number = self.current_row_number
        z = list( self.sheet.columns)
        for col in self.sheet.columns:
            try:
                #print(type(col))
                #print(type(col))
                self.sheet[col[0].column_letter + str(self.current_row_number)] = header[i]
            except Exception as e:
                print(e)
                return
            i = i + 1
            if i == len(header):
                break
        #self.sheet.insert_rows(r)
    def add_row(self,row):
        self.current_row_number = self.current_row_number + 1
        i = 1
        for col in self.sheet.columns:
            try:
                #print(type(col))
                #print(type(col))
                self.sheet[col[0].column_letter + str(self.current_row_number)] = row[i]
            except Exception as e:
                print(e)
                return
            i = i + 1
            if i == len(row):
                break
    def save_workbook(self, file_name):
        self.workbook.save(file_name + r'.xlsx')
